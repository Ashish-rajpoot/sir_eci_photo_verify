from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time

from selenium.common.exceptions import TimeoutException, StaleElementReferenceException



from config import PROFILE_PATH, USER_ID
from locators import LoginPageLocators, SearchPageLocators


def build_driver():
    options = Options()
    options.add_argument(f"--user-data-dir={PROFILE_PATH}")
    options.add_argument("--start-maximized")
    options.add_experimental_option("detach", True)

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    return driver


def is_login_page(wait):
    try:
        wait.until(EC.visibility_of_element_located(LoginPageLocators.USER_ID))
        return True
    except TimeoutException:
        return False


def fill_user_id_if_needed(wait):
    if is_login_page(wait):
        print("Login page detected.")

        user_box = wait.until(
            EC.visibility_of_element_located(LoginPageLocators.USER_ID)
        )
        user_box.clear()

        if USER_ID:
            user_box.send_keys(USER_ID)
            print("User ID filled from .env")
        else:
            print("ECI_USER_ID not found in .env")

        input("Complete captcha/OTP manually, then press Enter to continue...")
    else:
        print("Login page not detected. Possibly already logged in.")


def click_when_ready(wait, locator, name):
    element = wait.until(EC.element_to_be_clickable(locator))
    element.click()
    print(f"Clicked: {name}")


def read_excel_values(file_path, sheet_name="Sheet1"):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    values = []
    for row in ws.iter_rows(min_row=2, max_row=20, min_col=2, max_col=2, values_only=True):
        cell_value = row[0]
        if cell_value is not None and str(cell_value).strip():
            values.append(str(cell_value).strip())

    wb.close()
    return values



# def search_value(driver, wait, value):
    print(f"\n========== Processing value: {value} ==========")

    search_input = wait.until(
        EC.element_to_be_clickable(SearchPageLocators.SEARCH_INPUT)
    )

    search_input.click()
    time.sleep(0.5)

    search_input.clear()
    search_input.send_keys(value)
    time.sleep(1)

    # if dropdown option appears, click that instead of pressing ENTER
    option = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, f"//div[contains(@id,'option') and contains(.,'{value}')]")
        )
    )
    option.click()

    print(f"Selected option: {value}")

    try:
        click_when_ready(wait, SearchPageLocators.SEARCH_BUTTON, f"Search button for {value}")
        wait.until(EC.visibility_of_element_located(SearchPageLocators.TABLE))
        wait.until(EC.presence_of_all_elements_located(SearchPageLocators.TABLE_ROWS))
    except Exception as e:
        raise
def search_value(driver, wait, value):
    print(f"\n========== Processing value: {value} ==========")

    try:
        search_input = wait.until(
            EC.element_to_be_clickable(SearchPageLocators.SEARCH_INPUT)
        )

        search_input.click()
        time.sleep(0.5)

        search_input.clear()
        search_input.send_keys(value)
        time.sleep(1)

        # select dropdown option
        option = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, f"//div[contains(@id,'option') and contains(normalize-space(), \"{value}\")]")
            )
        )
        option.click()
        print(f"Selected option: {value}")

        click_when_ready(wait, SearchPageLocators.SEARCH_BUTTON, f"Search button for {value}")

        # wait for table
        wait.until(EC.presence_of_element_located(SearchPageLocators.TABLE))
        wait.until(EC.presence_of_all_elements_located(SearchPageLocators.TABLE_ROWS))

        rows = driver.find_elements(*SearchPageLocators.TABLE_ROWS)
        if not rows:
            print(f"{value} | table found but no rows | skipping")
            return False

        print(f"{value} | table loaded | row count = {len(rows)}")
        return True

    except TimeoutException:
        print(f"{value} | table/results not found | skipping this value")
        return False

    except Exception as e:
        print(f"{value} | unexpected error in search_value = {e}")
        return False
    
def get_table_rows(wait):
    wait.until(EC.visibility_of_element_located(SearchPageLocators.TABLE))
    rows = wait.until(EC.presence_of_all_elements_located(SearchPageLocators.TABLE_ROWS))
    return rows

def process_current_page_rows(driver, wait, value):
    # check table exists first
    tables = driver.find_elements(*SearchPageLocators.TABLE)

    if not tables:
        print(f"{value} | table not found | skipping and returning to main")
        return False

    rows = get_table_rows(wait)

    if not rows:
        print(f"{value} | no rows found | skipping and returning to main")
        return False

    for row_index in range(1, len(rows) + 1):
        status_xpath = f'//*[@id="tabWrapper"]/div[3]/div/table[1]/tbody/tr[{row_index}]/td[4]'
        action_xpath = f'//*[@id="tabWrapper"]/div[3]/div/table[1]/tbody/tr[{row_index}]/td[5]/i'

        try:
            status_text = wait.until(
                EC.visibility_of_element_located((By.XPATH, status_xpath))
            ).text.strip()
        except TimeoutException:
            print(f"{value} | row {row_index} | status cell not found | stopping page processing")
            return False

        print(f"{value} | row {row_index} | status = {status_text}")

        if status_text.lower() == "sir photo selected":
            continue

        elif status_text.lower() == "pending":
            try:
                action_icon = wait.until(
                    EC.element_to_be_clickable((By.XPATH, action_xpath))
                )
                action_icon.click()

                wait.until(EC.visibility_of_element_located(SearchPageLocators.POPUP))
                print(f"{value} | row {row_index} | popup opened")
                submit= wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="customModal"]/div/div[3]/div/button')
                    )
                )
                time.sleep(2)
                # UN-COMMENT IF YOU NEED TO AUTO CLICK
                # submit.click()
                
                # wait.until(EC.invisibility_of_element_located(SearchPageLocators.POPUP))
                # print(f"{value} | row {row_index} | popup closed")

                time.sleep(2)

            except TimeoutException:
                print(f"{value} | row {row_index} | popup/action not found | stopping page processing")
                return False

        else:
            print(f"{value} | row {row_index} | unexpected status = {status_text}")

    return True

# def process_current_page_rows(driver, wait, value):
#     tables = driver.find_elements(*SearchPageLocators.TABLE)

#     if not tables:
#         print(f"{value} | table not found | skipping and returning to main")
#         return False

#     rows = get_table_rows(wait)

#     for row_index in range(1, len(rows) + 1):
#         status_xpath = f'//*[@id="tabWrapper"]/div[3]/div/table[1]/tbody/tr[{row_index}]/td[4]'
#         action_xpath = f'//*[@id="tabWrapper"]/div[3]/div/table[1]/tbody/tr[{row_index}]/td[5]/i'

#         status_text = wait.until(
#             EC.visibility_of_element_located(("xpath", status_xpath))
#         ).text.strip()

#         print(f"{value} | row {row_index} | status = {status_text}")

#         if status_text.lower() == "sir photo selected":
#             continue

#         if status_text.lower() == "pending":
#             action_icon = wait.until(
#                 EC.element_to_be_clickable(("xpath", action_xpath))
#             )
#             action_icon.click()

#             wait.until(EC.visibility_of_element_located(SearchPageLocators.POPUP))
#             print(f"{value} | row {row_index} | popup opened")
#             time.sleep(2)
#             wait.until(EC.invisibility_of_element_located(SearchPageLocators.POPUP))
#             print(f"{value} | row {row_index} | popup Closed")
#             time.sleep(2)
#             # Stop here for now since you said popup is the next stage.
#             # Add popup actions later.
#             # input("Handle popup manually if needed, then press Enter to continue...")
#             # //*[@id="customModal"]/div/div[3]/div/button

#         else:
#             print(f"{value} | row {row_index} | unexpected status = {status_text}")


def has_next_page_enabled(driver, wait):
    next_btn = wait.until(
        EC.presence_of_element_located(SearchPageLocators.NEXT_PAGE_BUTTON)
    )
    disabled_attr = next_btn.get_attribute("disabled")
    return disabled_attr is None


def go_to_next_page(driver, wait):
    current_first_row = wait.until(
        EC.presence_of_element_located(("xpath", '//*[@id="tabWrapper"]/div[3]/div/table[1]/tbody/tr[1]'))
    )

    next_btn = wait.until(
        EC.element_to_be_clickable(SearchPageLocators.NEXT_PAGE_BUTTON)
    )
    next_btn.click()

    wait.until(EC.staleness_of(current_first_row))
    wait.until(EC.presence_of_all_elements_located(SearchPageLocators.TABLE_ROWS))


def process_all_pages_for_value(driver, wait, value):
    page_no = 1

    while True:
        print(f"{value} | processing page {page_no}")
        result = process_current_page_rows(driver, wait, value)

        if result is False:
            print(f"{value} | returning to main loop")
            return

        if has_next_page_enabled(driver, wait):
            print(f"{value} | moving to page {page_no + 1}")
            go_to_next_page(driver, wait)
            page_no += 1
        else:
            print(f"{value} | no more pages")
            wait.until(EC.visibility_of_element_located(SearchPageLocators.FIRST_PAGE_BUTTON))
            next_btn = wait.until(
                EC.element_to_be_clickable(SearchPageLocators.FIRST_PAGE_BUTTON)
            )
            next_btn.click()
            break

def reset_for_next_search(driver, wait):
    # close popup if present later when you know its close button
    # optional: scroll top
    driver.execute_script("window.scrollTo(0, 0);")

    # re-wait for input on current page
    wait.until(EC.visibility_of_element_located(SearchPageLocators.SEARCH_INPUT))

def force_enable_and_click(driver, wait, locator, name="element"):
    element = wait.until(EC.presence_of_element_located(locator))
    driver.execute_script("""
        arguments[0].removeAttribute('disabled');
        arguments[0].disabled = false;
    """, element)
    # driver.execute_script("arguments[0].click();", element)
    print(f"Force enabled and clicked: {name}")